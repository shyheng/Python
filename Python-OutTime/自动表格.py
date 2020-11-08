import openpyxl as xl
from  openpyxl.chart import BarChart, Reference


def process_workbook(filenme):
    wb =  xl.load_workbook(filenme)
    sheet = wb["Sheet1"]
    cell = sheet['a1']
    cell = sheet.cell(1,1)
    # print(cell.value)#编号，
    # print(sheet.max_row)#有多少行，4
    #
    # for row in range(1,sheet.max_row+1):
    #     print(row)#头表列
    #
    # for row in range(2,sheet.max_row+1):
    #     cell = sheet.cell(row,3)
    #     print(cell.value)#查找钱C列

    #添加D列的公式
    for row in range(2,sheet.max_row+1):
        cell = sheet.cell(row,3)
        diyi_qiang = cell.value * 0.9
        diyi_qiang_cell = sheet.cell(row, 4)
        diyi_qiang_cell.value = diyi_qiang
    #做表格
    values = Reference(sheet,
              min_row=2,
              max_row=sheet.max_row,
              min_col=4,
              max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart,"e2")

    wb.save(filenme)

